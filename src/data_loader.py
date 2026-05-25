"""Lecture des fichiers Excel HubSpot et filtrage par semaine."""

from __future__ import annotations

import datetime
import glob
import os

import pandas as pd
from openpyxl import load_workbook

from src.logger import get_logger

log = get_logger()

# Mapping des colonnes HubSpot → nom interne utilisé par l'app
COLUMN_MAP = {
    "Nom de l'entreprise": "NOM DU CLIENT",
    "Nom de la transaction": "NOM DU DOSSIER",
    "BU WMH": "MÉTIER PRINCIPAL",
    "Directeur conseil": "LEAD CONSEIL",
    "Propriétaire de la transaction": "LEAD PROJET",
    "Date de début de l'exploitation": "DATE DÉBUT",
}

# Patterns pour identifier le rôle de chaque fichier via le nom de sa feuille
# (case-insensitive, substring match). L'ordre importe : on teste les patterns
# les plus spécifiques en premier.
SHEET_PATTERNS = {
    "expo": "exploitation en co",  # Événements en cours d'exploitation cette semaine
    "gagné": "gagné j-7",          # Nouvelles signatures des 7 derniers jours
    "pipe": "reco à envoyer",      # Recos à envoyer cette semaine
}

# Valeurs HubSpot à traiter comme vides
EMPTY_VALUES = {"(Aucune valeur)", "Aucune valeur", ""}


def list_sheets(path: str) -> list[str]:
    """Retourne la liste des noms de feuilles d'un fichier Excel."""
    return _iter_workbook_sheets(path)


def _iter_workbook_sheets(path: str):
    """Retourne les noms de feuilles, robustement (liste vide si échec)."""
    try:
        wb = load_workbook(path, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


def find_excel_files(data_dir: str) -> dict:
    """
    Identifie les fichiers HubSpot dans le dossier en inspectant les noms
    de feuilles. Retourne {"expo": path, "gagné": path, "pipe": path}.
    """
    result: dict[str, str] = {}
    # Tri par mtime décroissant pour prendre le plus récent si plusieurs candidats
    candidates = sorted(
        glob.glob(os.path.join(data_dir, "*.xlsx")),
        key=os.path.getmtime, reverse=True,
    )

    for path in candidates:
        sheet_names = [s.lower() for s in _iter_workbook_sheets(path)]
        for role, pattern in SHEET_PATTERNS.items():
            if role in result:
                continue
            if any(pattern in name for name in sheet_names):
                result[role] = path
                log.info(f"Fichier {role!r} détecté : {os.path.basename(path)}")

    if not result:
        raise FileNotFoundError(
            f"Aucun fichier HubSpot reconnu dans {data_dir}. "
            "Attendu : une feuille contenant 'EXPLOITATION EN CO', 'Gagné J-7' ou 'reco à envoyer'."
        )

    return result


def _find_data_sheet(excel_path: str, label: str) -> str | int:
    """
    Trouve la feuille de données HubSpot (commençant par 'Entrée').
    Évite la feuille 'HubSpot Export Summary' qui a des colonnes différentes.
    """
    sheet_names = _iter_workbook_sheets(excel_path)
    for name in sheet_names:
        low = name.lower()
        if "entrée" in low or "entree" in low:
            return name
    # Fallback : seconde feuille (la première est le summary)
    return 1 if len(sheet_names) > 1 else 0


def validate_excel(excel_path: str, label: str = "Excel") -> None:
    """
    Valide qu'un fichier Excel est exploitable.
    Lève ValueError avec un message clair si ce n'est pas le cas.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Fichier {label} introuvable : {excel_path}")

    try:
        wb = load_workbook(excel_path, read_only=True)
    except Exception:
        raise ValueError(
            f"Le fichier {label} n'est pas un fichier Excel valide. "
            "Vérifiez qu'il s'agit bien d'un .xlsx."
        )

    sheet_names = wb.sheetnames
    wb.close()

    if not sheet_names:
        raise ValueError(f"Le fichier {label} ne contient aucune feuille.")

    log.info(f"Fichier {label} validé : {len(sheet_names)} feuilles — {', '.join(sheet_names)}")


def _apply_column_mapping(df: pd.DataFrame) -> pd.DataFrame:
    """Renomme les colonnes HubSpot vers le format interne."""
    rename = {k: v for k, v in COLUMN_MAP.items() if k in df.columns}
    if rename:
        log.info(f"Mapping colonnes appliqué : {list(rename.keys())}")
        df = df.rename(columns=rename)
    return df


def load_projects(excel_path: str, label: str = "Excel") -> pd.DataFrame:
    """Charge la feuille de données HubSpot."""
    validate_excel(excel_path, label)

    sheet = _find_data_sheet(excel_path, label)
    sheet_desc = f"'{sheet}'" if isinstance(sheet, str) else f"index {sheet}"

    try:
        df = pd.read_excel(excel_path, sheet_name=sheet)
    except Exception as e:
        raise ValueError(
            f"Impossible de lire la feuille {sheet_desc} du fichier {label} : {e}"
        )

    df = _apply_column_mapping(df)

    # Colonnes d'intérêt (toutes facultatives)
    cols_needed = [
        "NOM DU CLIENT", "NOM DU DOSSIER",
        "MÉTIER PRINCIPAL", "LEAD CONSEIL", "LEAD PROJET",
        "DATE DÉBUT",
    ]

    existing = [c for c in cols_needed if c in df.columns]
    df = df[existing].copy()

    # Normaliser la date si présente
    if "DATE DÉBUT" in df.columns:
        df["DATE DÉBUT"] = df["DATE DÉBUT"].apply(_normalize_date)

    # Lignes vides : pas de client ni de dossier → on jette
    if "NOM DU CLIENT" in df.columns and "NOM DU DOSSIER" in df.columns:
        mask_keep = df["NOM DU CLIENT"].apply(_is_non_empty) | df["NOM DU DOSSIER"].apply(_is_non_empty)
        df = df[mask_keep]

    log.info(f"Fichier {label} chargé ({sheet_desc}) : {len(df)} lignes")
    return df


def _is_non_empty(val) -> bool:
    if pd.isna(val):
        return False
    if isinstance(val, str) and val.strip() in EMPTY_VALUES:
        return False
    return True


def _normalize_date(val):
    """Convertit les valeurs de date Excel en datetime ou NaT."""
    if isinstance(val, (pd.Timestamp, datetime.datetime)):
        return pd.Timestamp(val)
    return pd.NaT


def get_week_bounds(reference_date: datetime.date = None):
    """Retourne (lundi, dimanche) de la semaine contenant reference_date."""
    if reference_date is None:
        reference_date = datetime.date.today()
    monday = reference_date - datetime.timedelta(days=reference_date.weekday())
    sunday = monday + datetime.timedelta(days=6)
    return (
        pd.Timestamp(monday),
        pd.Timestamp(sunday, hour=23, minute=59, second=59),
    )


def get_week_number(reference_date: datetime.date = None) -> int:
    """Retourne le numéro de semaine ISO."""
    if reference_date is None:
        reference_date = datetime.date.today()
    return reference_date.isocalendar()[1]


def filter_events_this_week(df: pd.DataFrame, reference_date: datetime.date = None) -> pd.DataFrame:
    """
    Pour le fichier 'Gagné J-7' : HubSpot a déjà filtré sur les 7 derniers
    jours de signature, on ne re-filtre donc pas par date d'exploitation.
    On trie simplement par date de début si elle est présente.
    """
    if "DATE DÉBUT" in df.columns:
        return df.sort_values("DATE DÉBUT", na_position="last")
    return df


def filter_expo_this_week(df: pd.DataFrame, reference_date: datetime.date = None) -> pd.DataFrame:
    """
    Pour le fichier 'EXPLOITATION EN COURS' : HubSpot a déjà filtré sur les
    événements qui jouent cette semaine, on retourne tout, trié par date.
    """
    if "DATE DÉBUT" in df.columns:
        return df.sort_values("DATE DÉBUT", na_position="last")
    return df


def filter_recos_this_week(df: pd.DataFrame, reference_date: datetime.date = None) -> pd.DataFrame:
    """
    Pour le fichier 'reco à envoyer cette semaine' : HubSpot a déjà filtré,
    on retourne tout. Pas de colonne date de rendu disponible.
    """
    return df


def _clean(val):
    """Vide si NaN, 0, '0' ou 'Aucune valeur'."""
    if pd.isna(val):
        return ""
    if isinstance(val, str):
        s = val.strip()
        if s in EMPTY_VALUES:
            return ""
        return s
    if val == 0 or val == "0":
        return ""
    return str(val).strip()


def format_expo_card(row: pd.Series) -> dict:
    """Formate une ligne 'EXPLOITATION EN COURS' en carte (mêmes champs qu'un événement)."""
    return format_event_card(row)


def format_event_card(row: pd.Series) -> dict:
    """Formate une ligne 'Gagné J-7' en carte."""
    def fmt_date(d):
        if pd.isna(d):
            return ""
        return d.strftime("%d/%m/%Y")

    return {
        "toolbox": "",
        "client": _clean(row.get("NOM DU CLIENT", "")),
        "dossier": _clean(row.get("NOM DU DOSSIER", "")),
        "dates": fmt_date(row.get("DATE DÉBUT")),
        "lead_conseil": _clean(row.get("LEAD CONSEIL", "")),
        "lead_projet": _clean(row.get("LEAD PROJET", "")),
        "squad": [],
        "lieu": "",
        "guests": "",
        "metier": _clean(row.get("MÉTIER PRINCIPAL", "")),
    }


def format_reco_card(row: pd.Series) -> dict:
    """Formate une ligne 'reco à envoyer' en carte (pas de colonne date)."""
    return {
        "toolbox": "",
        "client": _clean(row.get("NOM DU CLIENT", "")),
        "dossier": _clean(row.get("NOM DU DOSSIER", "")),
        "date_rendu": "",
        "lead_conseil": _clean(row.get("LEAD CONSEIL", "")),
        "lead_projet": _clean(row.get("LEAD PROJET", "")),
        "squad": [],
        "lieu": "",
        "guests": "",
        "metier": _clean(row.get("MÉTIER PRINCIPAL", "")),
    }
